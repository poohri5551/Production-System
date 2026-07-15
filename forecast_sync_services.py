import hashlib
import os
import re
import shutil
import tempfile
import time
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pymysql
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from part_services import normalize_part_no


SOURCE_SHEET = "FORECAST'18"
HEADER_ROW = 4
DATA_START_ROW = 5
PART_NO_COLUMN = 4
MONTH_START_COLUMN = 6  # F
TOTAL_HEADER_ROW = 3
# Unrelated forecast/stock sections follow Total (currently beginning at W3).
# Five years permits normal expansion while keeping a missing-boundary scan finite.
MAX_MONTH_COLUMNS = 60
EXCEL_FLOAT_SIGNIFICANT_DIGITS = 15
SOURCE_IDENTITY_STRATEGY = "source_sheet + source_row + normalized_part_no"
MONTH_LABEL_PATTERN = re.compile(
    r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-(\d{2})$",
    re.IGNORECASE,
)
MONTH_NUMBERS = {
    name.lower(): index
    for index, name in enumerate(
        ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"),
        start=1,
    )
}


class ForecastImportError(RuntimeError):
    pass


class ForecastStructuralDriftError(ForecastImportError):
    pass


class ForecastSourceUnavailableError(ForecastImportError):
    pass


def db_config():
    return {
        "host": os.environ.get("DB_HOST", "localhost"),
        "port": int(os.environ.get("DB_PORT", "3306")),
        "user": os.environ.get("DB_USER", "root"),
        "password": os.environ.get("DB_PASSWORD", ""),
        "database": os.environ.get("DB_NAME", "inventory_db"),
        "cursorclass": pymysql.cursors.DictCursor,
    }


def connect_database():
    return pymysql.connect(**db_config())


def is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def parse_part_no(value):
    if is_blank(value) or str(value).strip() == "-":
        return None, None, "blank"
    original = str(value).strip()
    normalized = normalize_part_no(original)
    if len(original) > 100 or not normalized or any(ord(character) < 32 for character in original):
        return original, normalized, "invalid"
    return original, normalized, None


def parse_forecast_quantity(value):
    if is_blank(value) or (isinstance(value, str) and value.strip() == "-"):
        return None, "blank"
    if isinstance(value, bool):
        return None, "boolean is not a forecast quantity"
    try:
        if isinstance(value, Decimal):
            parsed = value
        elif isinstance(value, int):
            parsed = Decimal(str(value))
        elif isinstance(value, float):
            parsed = Decimal(format(value, f".{EXCEL_FLOAT_SIGNIFICANT_DIGITS}g"))
        elif isinstance(value, str):
            parsed = Decimal(value.strip().replace(",", ""))
        else:
            return None, f"unsupported cell value type: {type(value).__name__}"
    except (InvalidOperation, ValueError):
        return None, "not a decimal number"
    if not parsed.is_finite():
        return None, "decimal is not finite"
    _, digits, exponent = parsed.as_tuple()
    scale = max(-exponent, 0)
    integer_digits = max(len(digits) - scale, 0) + max(exponent, 0)
    if scale > 10:
        return None, f"decimal scale {scale} exceeds database limit 10"
    if integer_digits > 20:
        return None, f"integer digits {integer_digits} exceed database limit 20"
    return parsed, None


def normalize_forecast_month(value):
    if isinstance(value, datetime):
        year, month = value.year, value.month
    elif isinstance(value, date):
        year, month = value.year, value.month
    elif isinstance(value, str):
        match = MONTH_LABEL_PATTERN.fullmatch(value.strip())
        if not match:
            raise ForecastImportError(f"Malformed or ambiguous FORECAST month header: {value!r}")
        month = MONTH_NUMBERS[match.group(1).lower()]
        buddhist_year = 2500 + int(match.group(2))
        year = buddhist_year - 543
        return date(year, month, 1)
    else:
        raise ForecastImportError(f"Unsupported FORECAST month header: {value!r}")

    if year >= 2400:
        year -= 543
    elif not 1900 <= year <= 2399:
        raise ForecastImportError(f"FORECAST month year is ambiguous: {value!r}")
    return date(year, month, 1)


def source_month_label(month):
    buddhist_year = month.year + 543
    return f"{month.strftime('%b')}-{buddhist_year % 100:02d}"


def source_identity(record):
    return (
        record["source_sheet"],
        int(record["source_row"]),
        record["normalized_part_no"],
    )


def find_duplicate_source_identities(records):
    occurrences = defaultdict(int)
    for record in records:
        occurrences[source_identity(record)] += 1
    return [
        {
            "source_sheet": identity[0],
            "source_row": identity[1],
            "normalized_part_no": identity[2],
            "record_count": count,
        }
        for identity, count in occurrences.items()
        if count > 1
    ]


def workbook_sha256(workbook_path):
    digest = hashlib.sha256()
    with Path(workbook_path).open("rb") as workbook_file:
        for chunk in iter(lambda: workbook_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def discover_month_layout(data_sheet):
    if str(data_sheet["D3"].value or "").strip().lower() != "part no.":
        raise ForecastImportError("D3 does not identify Part No.")

    scan_end_column = MONTH_START_COLUMN + MAX_MONTH_COLUMNS
    total_column = None
    for column in range(MONTH_START_COLUMN, scan_end_column + 1):
        boundary_value = data_sheet.cell(TOTAL_HEADER_ROW, column).value
        if boundary_value == "Total":
            total_column = column
            break
        if not is_blank(boundary_value):
            coordinate = f"{get_column_letter(column)}{TOTAL_HEADER_ROW}"
            raise ForecastImportError(
                "Exact Total boundary missing before downstream header "
                f"{boundary_value!r} at {coordinate}; CUSTOMER FORECAST layout drifted."
            )
    if total_column is None:
        scan_range = (
            f"{get_column_letter(MONTH_START_COLUMN)}{TOTAL_HEADER_ROW}:"
            f"{get_column_letter(scan_end_column)}{TOTAL_HEADER_ROW}"
        )
        raise ForecastImportError(
            f"Exact Total boundary not found within safe scan range {scan_range}; "
            "CUSTOMER FORECAST layout drifted."
        )
    if total_column == MONTH_START_COLUMN:
        raise ForecastImportError("CUSTOMER FORECAST monthly region contains no month columns.")

    months = []
    seen = set()
    previous_month = None
    for column in range(MONTH_START_COLUMN, total_column):
        raw_value = data_sheet.cell(HEADER_ROW, column).value
        if is_blank(raw_value):
            raise ForecastImportError(
                f"Blank month header at {get_column_letter(column)}{HEADER_ROW} before Total boundary."
            )
        month = normalize_forecast_month(raw_value)
        if month in seen:
            raise ForecastImportError(
                f"Duplicate month header at {get_column_letter(column)}{HEADER_ROW}."
            )
        if previous_month is not None and month <= previous_month:
            raise ForecastImportError(
                "CUSTOMER FORECAST month headers are not strictly increasing at "
                f"{get_column_letter(column)}{HEADER_ROW}."
            )
        seen.add(month)
        previous_month = month
        months.append({
            "column": column,
            "column_letter": get_column_letter(column),
            "source_label": source_month_label(month),
            "forecast_month": month,
            "month": month.strftime("%Y-%m"),
        })
    return months, total_column


def discover_month_columns(data_sheet):
    months, _total_column = discover_month_layout(data_sheet)
    return months


def parse_workbook(workbook_path, source_workbook_name=None):
    workbook_path = Path(workbook_path)
    source_workbook_name = source_workbook_name or workbook_path.name
    if not workbook_path.is_file():
        raise ForecastImportError(f"Workbook not found: {workbook_path}")
    with workbook_path.open("rb") as workbook_file:
        if workbook_file.read(4) != b"PK\x03\x04":
            raise ForecastImportError("Workbook is encrypted, malformed, or not a standard .xlsx file.")

    try:
        data_workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as error:
        raise ForecastImportError(
            "Workbook ZIP/XML is incomplete, malformed, or temporarily unreadable."
        ) from error
    formula_workbook = None
    try:
        worksheet_names = list(data_workbook.sheetnames)
        if SOURCE_SHEET not in worksheet_names:
            raise ForecastImportError(f"Required worksheet {SOURCE_SHEET!r} was not found.")
        try:
            formula_workbook = load_workbook(workbook_path, read_only=True, data_only=False)
        except Exception as error:
            raise ForecastImportError(
                "Workbook formula view is incomplete, malformed, or temporarily unreadable."
            ) from error
        data_sheet = data_workbook[SOURCE_SHEET]
        formula_sheet = formula_workbook[SOURCE_SHEET]
        months, total_column = discover_month_layout(data_sheet)
        total_column_letter = get_column_letter(total_column)
        final_month_column_letter = get_column_letter(total_column - 1)
        report = {
            "workbook": source_workbook_name,
            "workbook_sha256": workbook_sha256(workbook_path),
            "worksheet_names": worksheet_names,
            "selected_worksheet": SOURCE_SHEET,
            "header_row": HEADER_ROW,
            "part_no_column": "D",
            "month_region": f"F:{final_month_column_letter}",
            "month_boundary": f"{total_column_letter}{TOTAL_HEADER_ROW}=Total",
            "total_column": total_column,
            "total_column_letter": total_column_letter,
            "discovered_months": [
                {key: value for key, value in item.items() if key != "forecast_month"}
                for item in months
            ],
            "source_identity_strategy": SOURCE_IDENTITY_STRATEGY,
            "total_scanned_rows": 0,
            "blank_part_no_rows": [],
            "invalid_part_no_rows": [],
            "rows_without_monthly_quantity": [],
            "invalid_quantity_cells": [],
            "duplicate_part_nos": {},
            "true_duplicate_source_identities": [],
        }
        records = []
        part_occurrences = defaultdict(list)
        final_month_column = total_column - 1
        data_rows = data_sheet.iter_rows(min_row=DATA_START_ROW, max_col=final_month_column)
        formula_rows = formula_sheet.iter_rows(min_row=DATA_START_ROW, max_col=final_month_column)
        for source_row, (data_row, formula_row) in enumerate(zip(data_rows, formula_rows), DATA_START_ROW):
            report["total_scanned_rows"] += 1
            part_no, normalized_part_no, part_error = parse_part_no(data_row[PART_NO_COLUMN - 1].value)
            if part_error == "blank":
                report["blank_part_no_rows"].append(source_row)
                continue
            if part_error:
                report["invalid_part_no_rows"].append({"row": source_row, "value": part_no})
                continue
            monthly_values = []
            for month_info in months:
                index = month_info["column"] - 1
                quantity, quantity_error = parse_forecast_quantity(data_row[index].value)
                if quantity_error == "blank":
                    formula_cell = formula_row[index]
                    if formula_cell.data_type == "f":
                        report["invalid_quantity_cells"].append({
                            "row": source_row,
                            "column": month_info["column_letter"],
                            "month": month_info["month"],
                            "reason": "formula has no cached numeric value",
                        })
                    continue
                if quantity_error:
                    report["invalid_quantity_cells"].append({
                        "row": source_row,
                        "column": month_info["column_letter"],
                        "month": month_info["month"],
                        "reason": quantity_error,
                    })
                    continue
                monthly_values.append({
                    "forecast_month": month_info["forecast_month"],
                    "month": month_info["month"],
                    "source_label": month_info["source_label"],
                    "quantity": quantity,
                })
            if not monthly_values:
                report["rows_without_monthly_quantity"].append(source_row)
                continue
            part_occurrences[normalized_part_no].append(source_row)
            records.append({
                "part_no": part_no,
                "normalized_part_no": normalized_part_no,
                "source_workbook": source_workbook_name,
                "source_sheet": SOURCE_SHEET,
                "source_row": source_row,
                "monthly_values": monthly_values,
            })

        report["duplicate_part_nos"] = {
            value: rows for value, rows in part_occurrences.items() if len(rows) > 1
        }
        duplicates = find_duplicate_source_identities(records)
        report["true_duplicate_source_identities"] = duplicates
        report["true_duplicate_source_identity_rows"] = sum(item["record_count"] for item in duplicates)
        report["repeated_part_no_rows"] = sum(len(rows) for rows in report["duplicate_part_nos"].values())
        report["source_rows"] = len(records)
        report["monthly_quantity_values"] = sum(len(row["monthly_values"]) for row in records)
        report["invalid_rows"] = len(report["invalid_part_no_rows"]) + len(report["invalid_quantity_cells"])
        return records, report
    finally:
        data_workbook.close()
        if formula_workbook is not None:
            formula_workbook.close()


def table_exists(cursor, table_name):
    cursor.execute(
        "SELECT 1 FROM information_schema.TABLES "
        "WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME=%s LIMIT 1",
        (table_name,),
    )
    return cursor.fetchone() is not None


def load_database_state(cursor, require_monthly=True):
    if not table_exists(cursor, "forecast_entries"):
        raise ForecastImportError("forecast_entries table does not exist.")
    monthly_table_exists = table_exists(cursor, "forecast_monthly_values")
    if require_monthly and not monthly_table_exists:
        raise ForecastImportError("forecast_monthly_values table does not exist; migration 003 is required.")
    parts = {}
    if table_exists(cursor, "parts"):
        cursor.execute("SELECT id, normalized_part_no FROM parts WHERE deleted_at IS NULL")
        parts = {row["normalized_part_no"]: row["id"] for row in cursor.fetchall()}
    cursor.execute(
        "SELECT id, part_id, part_no, normalized_part_no, source_workbook, source_sheet, source_row, "
        "forecast_quantity, lot_count, lot_updated_at, lot_updated_by_user_id, lot_updated_by_username "
        "FROM forecast_entries"
    )
    parents = {source_identity(row): row for row in cursor.fetchall()}
    if monthly_table_exists:
        cursor.execute(
            "SELECT id, forecast_entry_id, forecast_month, source_label, quantity, lot_count, "
            "lot_updated_at, lot_updated_by_user_id, lot_updated_by_username "
            "FROM forecast_monthly_values"
        )
        monthly = {
            (int(row["forecast_entry_id"]), row["forecast_month"]): row
            for row in cursor.fetchall()
        }
    else:
        july = date(2026, 7, 1)
        monthly = {
            (int(row["id"]), july): {
                "forecast_entry_id": int(row["id"]),
                "forecast_month": july,
                "source_label": "Jul-69",
                "quantity": row["forecast_quantity"],
                "lot_count": row["lot_count"],
                "lot_updated_at": row["lot_updated_at"],
                "lot_updated_by_user_id": row["lot_updated_by_user_id"],
                "lot_updated_by_username": row["lot_updated_by_username"],
            }
            for row in parents.values()
        }
    return parts, parents, monthly


def load_month_registry(cursor, required=False):
    if not table_exists(cursor, "forecast_months"):
        if required:
            raise ForecastImportError("forecast_months table does not exist; migration 004 is required.")
        return {}
    cursor.execute(
        "SELECT forecast_month, source_label, is_active, first_seen_at, last_seen_at, "
        "deactivated_at FROM forecast_months"
    )
    return {row["forecast_month"]: row for row in cursor.fetchall()}


def canonical_discovered_months(records, discovered_months=None):
    canonical = {}
    if discovered_months is not None:
        for item in discovered_months:
            month = item.get("forecast_month")
            if month is None:
                try:
                    month = date.fromisoformat(f"{item['month']}-01")
                except (KeyError, TypeError, ValueError) as error:
                    raise ForecastImportError("Discovered FORECAST month metadata is invalid.") from error
            canonical[month] = item["source_label"]
        return canonical

    for record in records:
        for item in record.get("monthly_values", []):
            canonical[item["forecast_month"]] = item["source_label"]
    return canonical


def classify_month_registry(records, existing_months=None, discovered_months=None):
    existing_months = existing_months or {}
    discovered = canonical_discovered_months(records, discovered_months)
    inserts, reactivations, label_updates, deactivations, unchanged = [], [], [], [], []

    for forecast_month, source_label in sorted(discovered.items()):
        current = existing_months.get(forecast_month)
        prepared = {"forecast_month": forecast_month, "source_label": source_label}
        if current is None:
            inserts.append(prepared)
        elif not bool(current.get("is_active")):
            reactivations.append(prepared)
        elif current.get("source_label") != source_label:
            label_updates.append(prepared)
        else:
            unchanged.append(prepared)

    for forecast_month, current in sorted(existing_months.items()):
        if bool(current.get("is_active")) and forecast_month not in discovered:
            deactivations.append({
                "forecast_month": forecast_month,
                "source_label": current.get("source_label"),
            })

    return {
        "discovered_months": discovered,
        "month_inserts": inserts,
        "month_reactivations": reactivations,
        "month_label_updates": label_updates,
        "month_deactivations": deactivations,
        "month_unchanged": unchanged,
    }


def evaluate_structural_drift(records, existing_parents):
    duplicate_identities = find_duplicate_source_identities(records)
    existing_by_row = {int(row["source_row"]): row for row in existing_parents.values()}
    incoming_by_row = {int(row["source_row"]): row for row in records}
    row_conflicts = []
    for source_row in sorted(set(existing_by_row) & set(incoming_by_row)):
        before = existing_by_row[source_row]
        after = incoming_by_row[source_row]
        if before["normalized_part_no"] != after["normalized_part_no"]:
            row_conflicts.append({
                "source_row": source_row,
                "existing_part_no": before["normalized_part_no"],
                "incoming_part_no": after["normalized_part_no"],
            })
    existing_parts = {row["normalized_part_no"]: int(row["source_row"]) for row in existing_parents.values()}
    incoming_parts = {row["normalized_part_no"]: int(row["source_row"]) for row in records}
    moved = [
        {"part_no": part, "from_row": existing_parts[part], "to_row": incoming_parts[part]}
        for part in set(existing_parts) & set(incoming_parts)
        if abs(existing_parts[part] - incoming_parts[part]) > 5
    ]
    missing = [identity for identity in existing_parents if identity not in {source_identity(row) for row in records}]
    missing_limit = max(10, int(len(existing_parents) * 0.10))
    new_identities = [source_identity(row) for row in records if source_identity(row) not in existing_parents]
    new_limit = max(25, int(len(existing_parents) * 0.10))
    blocked_reasons = []
    if duplicate_identities:
        blocked_reasons.append("true duplicate source identities")
    if row_conflicts:
        blocked_reasons.append("existing source rows now contain different Part No.")
    if len(moved) > max(5, int(len(existing_parents) * 0.05)):
        blocked_reasons.append("large unexpected source-row movement")
    if len(missing) > missing_limit:
        blocked_reasons.append("abnormal disappearance of existing entries")
    if existing_parents and len(new_identities) > new_limit:
        blocked_reasons.append("abnormal reappearance/addition of many entries")
    return {
        "passed": not blocked_reasons,
        "blocked_reasons": blocked_reasons,
        "row_conflicts": row_conflicts,
        "large_movements": moved,
        "missing_existing_entries": len(missing),
        "missing_block_limit": missing_limit,
        "new_entry_count": len(new_identities),
        "new_entry_block_limit": new_limit,
        "duplicate_source_identities": duplicate_identities,
    }


def classify_sync(
    records,
    parts,
    existing_parents,
    existing_monthly,
    existing_months=None,
    discovered_months=None,
):
    drift = evaluate_structural_drift(records, existing_parents)
    parent_inserts, parent_updates, parent_unchanged = [], [], []
    monthly_inserts, monthly_updates, monthly_unchanged = [], [], []
    for record in records:
        prepared = dict(record)
        prepared["part_id"] = parts.get(record["normalized_part_no"])
        existing = existing_parents.get(source_identity(record))
        if existing is None:
            parent_inserts.append(prepared)
            for value in record["monthly_values"]:
                monthly_inserts.append((prepared, value))
            continue
        prepared["id"] = int(existing["id"])
        comparable = ("part_id", "part_no", "source_workbook")
        (parent_updates if any(existing.get(field) != prepared.get(field) for field in comparable) else parent_unchanged).append(prepared)
        for value in record["monthly_values"]:
            current = existing_monthly.get((prepared["id"], value["forecast_month"]))
            pair = (prepared, value)
            if current is None:
                monthly_inserts.append(pair)
            elif Decimal(str(current["quantity"])) != value["quantity"] or current.get("source_label") != value["source_label"]:
                monthly_updates.append(pair)
            else:
                monthly_unchanged.append(pair)
    incoming = {source_identity(row) for row in records}
    stale = [row for identity, row in existing_parents.items() if identity not in incoming]
    plan = {
        "parent_inserts": parent_inserts,
        "parent_updates": parent_updates,
        "parent_unchanged": parent_unchanged,
        "monthly_inserts": monthly_inserts,
        "monthly_updates": monthly_updates,
        "monthly_unchanged": monthly_unchanged,
        "stale_parents": stale,
        "structural_drift": drift,
    }
    plan.update(classify_month_registry(records, existing_months, discovered_months))
    return plan


def apply_sync(cursor, plan):
    if not plan["structural_drift"]["passed"]:
        raise ForecastStructuralDriftError(
            "Structural drift blocked: " + "; ".join(plan["structural_drift"]["blocked_reasons"])
        )
    for record in plan["parent_inserts"]:
        cursor.execute(
            "INSERT INTO forecast_entries "
            "(part_id, part_no, normalized_part_no, source_workbook, source_sheet, source_row) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (record["part_id"], record["part_no"], record["normalized_part_no"],
             record["source_workbook"], record["source_sheet"], record["source_row"]),
        )
        record["id"] = cursor.lastrowid
    for record in plan["parent_updates"]:
        cursor.execute(
            "UPDATE forecast_entries SET part_id=%s, part_no=%s, source_workbook=%s WHERE id=%s",
            (record["part_id"], record["part_no"], record["source_workbook"], record["id"]),
        )
    for record, value in plan["monthly_inserts"]:
        cursor.execute(
            "INSERT INTO forecast_monthly_values "
            "(forecast_entry_id, forecast_month, source_label, quantity) VALUES (%s,%s,%s,%s)",
            (record["id"], value["forecast_month"], value["source_label"], value["quantity"]),
        )
    for record, value in plan["monthly_updates"]:
        cursor.execute(
            "UPDATE forecast_monthly_values SET source_label=%s, quantity=%s "
            "WHERE forecast_entry_id=%s AND forecast_month=%s",
            (value["source_label"], value["quantity"], record["id"], value["forecast_month"]),
        )
    for item in plan.get("month_inserts", []):
        cursor.execute(
            "INSERT INTO forecast_months "
            "(forecast_month, source_label, is_active, first_seen_at, last_seen_at, deactivated_at) "
            "VALUES (%s,%s,1,NOW(),NOW(),NULL)",
            (item["forecast_month"], item["source_label"]),
        )
    for item in plan.get("month_reactivations", []):
        cursor.execute(
            "UPDATE forecast_months SET source_label=%s, is_active=1, last_seen_at=NOW(), "
            "deactivated_at=NULL WHERE forecast_month=%s AND is_active=0",
            (item["source_label"], item["forecast_month"]),
        )
    for item in plan.get("month_label_updates", []):
        cursor.execute(
            "UPDATE forecast_months SET source_label=%s, last_seen_at=NOW() "
            "WHERE forecast_month=%s AND is_active=1 AND source_label<>%s",
            (item["source_label"], item["forecast_month"], item["source_label"]),
        )
    for item in plan.get("month_deactivations", []):
        cursor.execute(
            "UPDATE forecast_months SET is_active=0, deactivated_at=NOW() "
            "WHERE forecast_month=%s AND is_active=1",
            (item["forecast_month"],),
        )


def month_report(plan):
    discovered = plan.get("discovered_months", {})
    return {
        "months_discovered": len(discovered),
        "months_active": len(discovered),
        "months_activated": len(plan.get("month_inserts", [])),
        "months_reactivated": len(plan.get("month_reactivations", [])),
        "months_deactivated": len(plan.get("month_deactivations", [])),
        "months_unchanged": len(plan.get("month_unchanged", [])),
    }


def planned_database_writes(plan):
    return sum(
        len(plan.get(key, []))
        for key in (
            "parent_inserts",
            "parent_updates",
            "monthly_inserts",
            "monthly_updates",
            "month_inserts",
            "month_reactivations",
            "month_label_updates",
            "month_deactivations",
        )
    )


def sync_workbook(workbook_path, apply=False, connection_factory=connect_database):
    records, report = parse_workbook(workbook_path)
    if apply and report["invalid_rows"]:
        raise ForecastImportError(
            f"Apply blocked: workbook contains {report['invalid_rows']} invalid Part/month cells."
        )
    connection = connection_factory()
    try:
        with connection.cursor() as cursor:
            parts, parents, monthly = load_database_state(cursor, require_monthly=apply)
            existing_months = load_month_registry(cursor, required=apply)
            plan = classify_sync(
                records,
                parts,
                parents,
                monthly,
                existing_months=existing_months,
                discovered_months=report["discovered_months"],
            )
            if apply:
                apply_sync(cursor, plan)
        if apply:
            connection.commit()
        report.update({
            "parent_inserts": len(plan["parent_inserts"]),
            "parent_updates": len(plan["parent_updates"]),
            "parent_unchanged": len(plan["parent_unchanged"]),
            "monthly_inserts": len(plan["monthly_inserts"]),
            "monthly_updates": len(plan["monthly_updates"]),
            "monthly_unchanged": len(plan["monthly_unchanged"]),
            "stale_parents": len(plan["stale_parents"]),
            "structural_drift": plan["structural_drift"],
            "database_writes": planned_database_writes(plan) if apply else 0,
            "mode": "APPLY" if apply else "DRY-RUN",
        })
        report.update(month_report(plan))
        return report
    except Exception:
        if apply:
            connection.rollback()
        raise
    finally:
        connection.close()


def stable_snapshot(source_path, stability_seconds=3.0):
    source_path = Path(source_path)
    before = source_path.stat()
    time.sleep(max(0, stability_seconds))
    after = source_path.stat()
    if (before.st_size, before.st_mtime_ns) != (after.st_size, after.st_mtime_ns):
        raise ForecastImportError("Workbook is still changing; retry later.")
    temporary = tempfile.NamedTemporaryFile(prefix="forecast-", suffix=".xlsx", delete=False)
    snapshot_path = Path(temporary.name)
    try:
        with source_path.open("rb") as source, temporary:
            shutil.copyfileobj(source, temporary)
        final = source_path.stat()
        if (after.st_size, after.st_mtime_ns) != (final.st_size, final.st_mtime_ns):
            raise ForecastImportError("Workbook changed during snapshot; retry later.")
        if snapshot_path.stat().st_size != final.st_size:
            raise ForecastImportError("Workbook snapshot size mismatch; retry later.")
        return snapshot_path
    except Exception:
        snapshot_path.unlink(missing_ok=True)
        raise
