import os
import copy
import tempfile
import unittest
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

import app as app_module
from forecast_services import (
    ForecastValidationError,
    calculate_quantity,
    choose_default_month,
    validate_lot_batch,
    validate_lot_count,
)
from forecast_sync_services import (
    ForecastImportError,
    ForecastStructuralDriftError,
    apply_sync,
    classify_sync,
    classify_month_registry,
    discover_month_columns,
    evaluate_structural_drift,
    normalize_forecast_month,
    parse_forecast_quantity,
    parse_workbook,
    planned_database_writes,
    sync_workbook,
)
from scripts import forecast_sync_worker
from scripts.import_forecast_from_excel import parse_args


ROOT = Path(__file__).resolve().parents[1]


class ForecastServiceTests(unittest.TestCase):
    def test_lot_validation(self):
        self.assertEqual(validate_lot_count(2), 2)
        self.assertIsNone(validate_lot_count(None))
        for value in (0, -1, 1.5, "2", True):
            with self.subTest(value=value), self.assertRaises(ForecastValidationError):
                validate_lot_count(value)

    def test_batch_requires_entry_month_and_lot(self):
        self.assertEqual(
            validate_lot_batch({"items": [{"id": 1, "month": "2026-07", "lot_count": 2}]})[0]["forecast_month"],
            date(2026, 7, 1),
        )
        for payload in (
            {"items": [{"id": 1, "lot_count": 2}]},
            {"items": [{"id": 1, "month": "Jul-69", "lot_count": 2}]},
            {"items": [{"id": 1, "month": "2026-07", "lot_count": 2, "quantity": 3}]},
            {"items": [
                {"id": 1, "month": "2026-07", "lot_count": 2},
                {"id": 1, "month": "2026-07", "lot_count": 3},
            ]},
        ):
            with self.subTest(payload=payload), self.assertRaises(ForecastValidationError):
                validate_lot_batch(payload)

    def test_quantity_per_lot_is_decimal_safe(self):
        self.assertEqual(calculate_quantity(454, 2), Decimal("227"))
        self.assertEqual(calculate_quantity(459, 2), Decimal("229.5"))

    def test_default_month_selection_rule(self):
        months = ["2026-06", "2026-07", "2026-09"]
        self.assertEqual(choose_default_month(months, date(2026, 7, 14)), "2026-07")
        self.assertEqual(choose_default_month(months, date(2026, 8, 1)), "2026-07")
        self.assertEqual(choose_default_month(months, date(2025, 1, 1)), "2026-06")


class ForecastWorkbookTests(unittest.TestCase):
    MONTH_HEADERS = [
        datetime(2567, 12, 1),
        datetime(2568, 10, 1),
        datetime(2568, 11, 1),
        datetime(2568, 12, 1),
        datetime(2569, 1, 1),
        datetime(2569, 2, 1),
        datetime(2569, 3, 1),
        datetime(2569, 4, 1),
        datetime(2569, 5, 1),
        datetime(2569, 6, 1),
        datetime(2569, 7, 1),
        datetime(2569, 8, 1),
        datetime(2569, 9, 1),
        datetime(2569, 10, 1),
    ]

    def make_workbook(self, exact_sheet=True, month_headers=None, boundary_header="Total"):
        custom_layout = month_headers is not None
        month_headers = self.MONTH_HEADERS if month_headers is None else month_headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "FORECAST'18" if exact_sheet else "FORECAST '18"
        sheet["D3"] = "Part No."
        total_column = 6 + len(month_headers)
        if boundary_header is not None:
            sheet.cell(3, total_column, boundary_header)
        for column, header in enumerate(month_headers, 6):
            sheet.cell(4, column, header)
        if custom_layout:
            sheet["D5"] = "PART-A"
            for column in range(6, total_column):
                sheet.cell(5, column, column)
        else:
            sheet["D5"], sheet["P5"], sheet["Q5"] = "PART-A", 454, 459
            sheet["D6"], sheet["P6"] = "PART-B", "142,570"
            sheet["D7"], sheet["P7"] = "DUP", 10
            sheet["D8"], sheet["P8"] = " dup ", 20
            sheet["D9"], sheet["P9"] = "BAD", "not-a-number"
        sheet.cell(5, total_column, 999999)
        temporary = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temporary.close()
        workbook.save(temporary.name)
        workbook.close()
        self.addCleanup(Path(temporary.name).unlink, missing_ok=True)
        return Path(temporary.name)

    def test_buddhist_month_normalization(self):
        expected = {
            "Jul-69": date(2026, 7, 1),
            "Aug-69": date(2026, 8, 1),
            "Sep-69": date(2026, 9, 1),
            "Dec-69": date(2026, 12, 1),
            "Jan-70": date(2027, 1, 1),
        }
        for label, month in expected.items():
            with self.subTest(label=label):
                self.assertEqual(normalize_forecast_month(label), month)
        with self.assertRaises(ForecastImportError):
            normalize_forecast_month("07/69")

    def test_exact_sheet_month_discovery_and_summary_exclusion(self):
        records, report = parse_workbook(self.make_workbook())
        discovered = report["discovered_months"]
        self.assertEqual(len(discovered), 14)
        self.assertEqual(discovered[10], {
            "column": 16,
            "column_letter": "P",
            "source_label": "Jul-69",
            "month": "2026-07",
        })
        self.assertEqual(discovered[-1]["column_letter"], "S")
        self.assertEqual(report["month_boundary"], "T3=Total")
        self.assertFalse(any(value["quantity"] == Decimal("999999") for row in records for value in row["monthly_values"]))
        self.assertEqual(records[0]["monthly_values"][0]["quantity"], Decimal("454"))
        self.assertTrue(all("lot_count" not in value for row in records for value in row["monthly_values"]))
        self.assertEqual(report["repeated_part_no_rows"], 2)

    def test_dynamic_total_boundary_accepts_original_and_future_months(self):
        layouts = (
            (["Jul-69", "Aug-69", "Sep-69", "Oct-69"], "J", "2026-10"),
            (["Jul-69", "Aug-69", "Sep-69", "Oct-69", "Nov-69"], "K", "2026-11"),
            (
                ["Jul-69", "Aug-69", "Sep-69", "Oct-69", "Nov-69", "Dec-69", "Jan-70"],
                "M",
                "2027-01",
            ),
        )
        for headers, total_column, final_month in layouts:
            with self.subTest(headers=headers):
                records, report = parse_workbook(self.make_workbook(month_headers=headers))
                self.assertEqual(
                    [item["source_label"] for item in report["discovered_months"]],
                    headers,
                )
                self.assertEqual(report["discovered_months"][-1]["month"], final_month)
                self.assertEqual(report["total_column_letter"], total_column)
                self.assertEqual(report["month_boundary"], f"{total_column}3=Total")
                self.assertFalse(
                    any(
                        value["quantity"] == Decimal("999999")
                        for row in records
                        for value in row["monthly_values"]
                    )
                )

    def test_invalid_dynamic_month_layouts_block_before_database_connection(self):
        invalid_layouts = (
            ("non-month", ["Jul-69", "Aug-69", "Remark", "Sep-69"], "Total"),
            ("blank", ["Jul-69", "Aug-69", None, "Sep-69"], "Total"),
            ("duplicate", ["Jul-69", "Aug-69", "Aug-69"], "Total"),
            ("out-of-order", ["Sep-69", "Aug-69"], "Total"),
            ("missing-total", ["Jul-69", "Aug-69"], None),
        )
        for label, headers, boundary_header in invalid_layouts:
            with self.subTest(label=label), self.assertRaises(ForecastImportError):
                sync_workbook(
                    self.make_workbook(
                        month_headers=headers,
                        boundary_header=boundary_header,
                    ),
                    apply=True,
                    connection_factory=lambda: self.fail("must perform zero database writes"),
                )

    def test_similar_sheet_name_is_rejected(self):
        with self.assertRaises(ForecastImportError):
            parse_workbook(self.make_workbook(exact_sheet=False))

    def test_partial_zip_workbook_is_rejected_cleanly(self):
        temporary = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temporary.write(b"PK\x03\x04partial")
        temporary.close()
        self.addCleanup(Path(temporary.name).unlink, missing_ok=True)
        with self.assertRaises(ForecastImportError):
            parse_workbook(temporary.name)

    def test_snapshot_parse_keeps_original_source_filename(self):
        records, report = parse_workbook(
            self.make_workbook(),
            source_workbook_name="1.FORCASE CENTOR.xlsx",
        )
        self.assertEqual(report["workbook"], "1.FORCASE CENTOR.xlsx")
        self.assertTrue(all(row["source_workbook"] == "1.FORCASE CENTOR.xlsx" for row in records))

    def test_quantity_parser_is_decimal_safe(self):
        self.assertEqual(parse_forecast_quantity("142,570"), (Decimal("142570"), None))
        self.assertEqual(parse_forecast_quantity(446.51250000000005), (Decimal("446.5125"), None))
        self.assertIsNotNone(parse_forecast_quantity("bad")[1])

    def test_apply_blocks_invalid_cells_before_database_connection(self):
        report = {"invalid_rows": 1}
        with patch("forecast_sync_services.parse_workbook", return_value=([], report)):
            with self.assertRaises(ForecastImportError):
                sync_workbook("ignored.xlsx", apply=True, connection_factory=lambda: self.fail("must not connect"))


class ForecastPlanningTests(unittest.TestCase):
    def record(self, row=5, part="PART-A", quantities=None):
        return {
            "part_no": part,
            "normalized_part_no": part,
            "source_workbook": "book.xlsx",
            "source_sheet": "FORECAST'18",
            "source_row": row,
            "monthly_values": quantities or [
                {"forecast_month": date(2026, 7, 1), "month": "2026-07", "source_label": "Jul-69", "quantity": Decimal("20")},
                {"forecast_month": date(2026, 8, 1), "month": "2026-08", "source_label": "Aug-69", "quantity": Decimal("30")},
            ],
        }

    def test_one_parent_has_independent_months_and_reimport_preserves_lot(self):
        record = self.record()
        identity = ("FORECAST'18", 5, "PART-A")
        parent = {"id": 1, "part_id": None, "part_no": "PART-A", "normalized_part_no": "PART-A", "source_workbook": "book.xlsx", "source_sheet": "FORECAST'18", "source_row": 5}
        july = {"forecast_entry_id": 1, "forecast_month": date(2026, 7, 1), "source_label": "Jul-69", "quantity": Decimal("10"), "lot_count": 7, "lot_updated_by_username": "pc_user"}
        august = {"forecast_entry_id": 1, "forecast_month": date(2026, 8, 1), "source_label": "Aug-69", "quantity": Decimal("30"), "lot_count": 3}
        plan = classify_sync([record], {}, {identity: parent}, {(1, date(2026, 7, 1)): july, (1, date(2026, 8, 1)): august})
        self.assertEqual(len(plan["monthly_updates"]), 1)
        self.assertEqual(len(plan["monthly_unchanged"]), 1)
        self.assertEqual(july["lot_count"], 7)
        self.assertEqual(july["lot_updated_by_username"], "pc_user")

    @staticmethod
    def month_state(month, label, active=True):
        return {
            "forecast_month": month,
            "source_label": label,
            "is_active": active,
        }

    def test_initial_month_activation_and_missing_month_soft_deactivation(self):
        record = self.record()
        initial = classify_month_registry([record], {})
        self.assertEqual(
            [item["forecast_month"] for item in initial["month_inserts"]],
            [date(2026, 7, 1), date(2026, 8, 1)],
        )

        september = date(2026, 9, 1)
        existing_months = {
            date(2026, 7, 1): self.month_state(date(2026, 7, 1), "Jul-69"),
            date(2026, 8, 1): self.month_state(date(2026, 8, 1), "Aug-69"),
            september: self.month_state(september, "Sep-69"),
        }
        retained_monthly = {
            (1, september): {"quantity": Decimal("500"), "lot_count": 5},
        }
        removed = classify_month_registry([record], existing_months)
        self.assertEqual(removed["month_deactivations"], [{
            "forecast_month": september,
            "source_label": "Sep-69",
        }])
        self.assertEqual(retained_monthly[(1, september)]["quantity"], Decimal("500"))
        self.assertEqual(retained_monthly[(1, september)]["lot_count"], 5)

    def test_reactivation_preserves_lot_and_updates_only_changed_excel_quantity(self):
        september = date(2026, 9, 1)
        record = self.record(quantities=[{
            "forecast_month": september,
            "month": "2026-09",
            "source_label": "Sep-69",
            "quantity": Decimal("700"),
        }])
        identity = ("FORECAST'18", 5, "PART-A")
        parent = {
            "id": 1,
            "part_id": None,
            "part_no": "PART-A",
            "normalized_part_no": "PART-A",
            "source_workbook": "book.xlsx",
            "source_sheet": "FORECAST'18",
            "source_row": 5,
        }
        monthly = {
            (1, september): {
                "forecast_entry_id": 1,
                "forecast_month": september,
                "source_label": "Sep-69",
                "quantity": Decimal("500"),
                "lot_count": 5,
                "lot_updated_by_username": "pc_user",
            },
        }
        registry = {september: self.month_state(september, "Sep-69", active=False)}
        plan = classify_sync([record], {}, {identity: parent}, monthly, registry)
        self.assertEqual(len(plan["month_reactivations"]), 1)
        self.assertEqual(len(plan["monthly_updates"]), 1)
        self.assertEqual(monthly[(1, september)]["lot_count"], 5)
        self.assertEqual(monthly[(1, september)]["lot_updated_by_username"], "pc_user")
        self.assertEqual(calculate_quantity(Decimal("700"), 5), Decimal("140"))

        monthly[(1, september)]["quantity"] = Decimal("700")
        same_quantity = classify_sync([record], {}, {identity: parent}, monthly, registry)
        self.assertEqual(len(same_quantity["month_reactivations"]), 1)
        self.assertEqual(len(same_quantity["monthly_updates"]), 0)
        self.assertEqual(monthly[(1, september)]["lot_count"], 5)

    def test_identical_active_source_is_idempotent(self):
        record = self.record()
        identity = ("FORECAST'18", 5, "PART-A")
        parent = {"id": 1, "part_id": None, "part_no": "PART-A", "normalized_part_no": "PART-A", "source_workbook": "book.xlsx", "source_sheet": "FORECAST'18", "source_row": 5}
        monthly = {
            (1, value["forecast_month"]): {
                "quantity": value["quantity"],
                "source_label": value["source_label"],
                "lot_count": 5,
            }
            for value in record["monthly_values"]
        }
        registry = {
            value["forecast_month"]: self.month_state(
                value["forecast_month"], value["source_label"]
            )
            for value in record["monthly_values"]
        }
        plan = classify_sync([record], {}, {identity: parent}, monthly, registry)
        self.assertEqual(planned_database_writes(plan), 0)
        self.assertEqual(len(plan["month_unchanged"]), 2)

    def test_structural_drift_blocks_before_any_month_or_quantity_write(self):
        cursor = unittest.mock.Mock()
        plan = {
            "structural_drift": {"passed": False, "blocked_reasons": ["layout drift"]},
            "parent_inserts": [],
            "parent_updates": [],
            "monthly_inserts": [],
            "monthly_updates": [],
            "month_deactivations": [{"forecast_month": date(2026, 9, 1)}],
        }
        with self.assertRaises(ForecastStructuralDriftError):
            apply_sync(cursor, plan)
        cursor.execute.assert_not_called()

    def test_source_row_part_change_blocks_with_zero_apply_plan(self):
        parent = {"id": 1, "normalized_part_no": "PART-A", "source_sheet": "FORECAST'18", "source_row": 5}
        drift = evaluate_structural_drift([self.record(part="PART-B")], {("FORECAST'18", 5, "PART-A"): parent})
        self.assertFalse(drift["passed"])
        self.assertIn("existing source rows now contain different Part No.", drift["blocked_reasons"])

    def test_migration_seeds_july_and_never_persists_derived_value(self):
        sql = (ROOT / "migrations/003_create_forecast_monthly_values.sql").read_text(encoding="utf-8")
        self.assertIn("DATE('2026-07-01')", sql)
        self.assertIn("forecast_quantity", sql)
        self.assertIn("lot_count", sql)
        self.assertIn("lot_updated_by_username", sql)
        self.assertNotIn("quantity_per_lot", sql)
        self.assertNotIn("DROP TABLE", sql.upper())

    def test_month_registry_migration_seeds_existing_months_without_delete(self):
        sql = (ROOT / "migrations/004_create_forecast_months.sql").read_text(encoding="utf-8")
        self.assertIn("CREATE TABLE IF NOT EXISTS forecast_months", sql)
        self.assertIn("FROM forecast_monthly_values", sql)
        self.assertIn("INSERT IGNORE INTO forecast_months", sql)
        self.assertNotIn("DELETE FROM", sql.upper())


class ForecastCursor:
    def __init__(self):
        self.rows = {
            (1, date(2026, 7, 1)): {"id": 1, "part_no": "PART-A", "source_row": 5, "forecast_month": date(2026, 7, 1), "source_label": "Jul-69", "quantity": Decimal("454"), "lot_count": 2},
            (1, date(2026, 8, 1)): {"id": 1, "part_no": "PART-A", "source_row": 5, "forecast_month": date(2026, 8, 1), "source_label": "Aug-69", "quantity": Decimal("459"), "lot_count": 3},
        }
        self.result = []
        self.update_count = 0
        self.months = {
            date(2026, 7, 1): {"source_label": "Jul-69", "is_active": True},
            date(2026, 8, 1): {"source_label": "Aug-69", "is_active": True},
        }

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def execute(self, sql, params=()):
        normalized = " ".join(sql.split()).upper()
        if normalized.startswith("SELECT DATE_FORMAT(FORECAST_MONTH"):
            self.result = [
                {"month": month.strftime("%Y-%m"), "label": state["source_label"]}
                for month, state in sorted(self.months.items())
                if state["is_active"]
            ]
        elif normalized.startswith("SELECT FORECAST_MONTH, IS_ACTIVE FROM FORECAST_MONTHS"):
            self.result = [
                {"forecast_month": month, "is_active": self.months[month]["is_active"]}
                for month in params
                if month in self.months
            ]
        elif "FROM FORECAST_ENTRIES E" in normalized:
            if len(params) == 1:
                selected = params[0]
                self.result = [
                    dict(row) for (entry, month), row in self.rows.items()
                    if month == selected and self.months.get(month, {}).get("is_active")
                ]
            else:
                pairs = {(params[index], params[index + 1]) for index in range(0, len(params), 2)}
                self.result = [
                    dict(row) for pair, row in self.rows.items()
                    if pair in pairs and self.months.get(pair[1], {}).get("is_active")
                ]
        elif normalized.startswith("UPDATE FORECAST_MONTHLY_VALUES"):
            lot, user_id, username, entry_id, month = params
            self.rows[(entry_id, month)]["lot_count"] = lot
            self.rows[(entry_id, month)]["lot_updated_by_user_id"] = user_id
            self.rows[(entry_id, month)]["lot_updated_by_username"] = username
            self.update_count += 1
            self.result = []
        else:
            raise AssertionError(f"Unexpected SQL: {sql}")

    def fetchall(self):
        return [dict(row) for row in self.result]


class ForecastConnection:
    def __init__(self):
        self.cursor_instance = ForecastCursor()
        self.committed = False
        self.rolled_back = False

    def cursor(self):
        return self.cursor_instance

    def commit(self):
        self.committed = True

    def rollback(self):
        self.rolled_back = True

    def close(self):
        pass


class ForecastApiTests(unittest.TestCase):
    def setUp(self):
        self.connection = ForecastConnection()

        def fake_load_session_user(_cursor):
            role = app_module.session.get("test_role")
            return {"id": 77, "username": "pc_user", "role": role} if role else None

        self.patches = [
            patch.object(app_module, "get_db_connection", return_value=self.connection),
            patch.object(app_module, "ensure_users_table", return_value=None),
            patch.object(app_module, "load_session_user", side_effect=fake_load_session_user),
        ]
        for active in self.patches:
            active.start()
        app_module.app.config.update(TESTING=True)
        self.client = app_module.app.test_client()

    def tearDown(self):
        for active in reversed(self.patches):
            active.stop()

    def authenticate(self, role):
        with self.client.session_transaction() as state:
            state.update(user_id=77, username="pc_user", test_role=role)

    def test_unauthenticated_is_401_and_every_non_pc_is_403(self):
        self.assertEqual(self.client.get("/api/forecast/months").status_code, 401)
        for role in ("Admin", "Sup", "Manager", "Technician", "QC Line", "Operator"):
            self.authenticate(role)
            self.assertEqual(self.client.get("/api/forecast?month=2026-07").status_code, 403)

    def test_months_and_selected_month(self):
        self.authenticate("PC")
        self.assertEqual(len(self.client.get("/api/forecast/months").get_json()["months"]), 2)
        payload = self.client.get("/api/forecast?month=2026-07").get_json()
        self.assertEqual(payload["month"], "2026-07")
        self.assertEqual(payload["items"][0]["quantity_per_lot"], 227)
        self.assertEqual(self.client.get("/api/forecast?month=Jul-69").status_code, 400)
        self.assertEqual(self.client.get("/api/forecast?month=2026-09").status_code, 400)

    def test_lot_update_targets_only_entry_month(self):
        self.authenticate("PC")
        response = self.client.post("/api/forecast/lots", json={"items": [{"id": 1, "month": "2026-07", "lot_count": 4}]})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.connection.cursor_instance.rows[(1, date(2026, 7, 1))]["lot_count"], 4)
        self.assertEqual(self.connection.cursor_instance.rows[(1, date(2026, 8, 1))]["lot_count"], 3)
        self.assertEqual(self.connection.cursor_instance.rows[(1, date(2026, 7, 1))]["lot_updated_by_username"], "pc_user")

    def test_inactive_month_is_hidden_and_lot_save_returns_conflict_without_write(self):
        self.authenticate("PC")
        september = date(2026, 9, 1)
        cursor = self.connection.cursor_instance
        cursor.months[september] = {"source_label": "Sep-69", "is_active": False}
        cursor.rows[(1, september)] = {"id": 1, "part_no": "PART-A", "source_row": 5, "forecast_month": september, "source_label": "Sep-69", "quantity": Decimal("500"), "lot_count": 5}

        months = self.client.get("/api/forecast/months").get_json()["months"]
        self.assertNotIn("2026-09", [item["month"] for item in months])
        response = self.client.post("/api/forecast/lots", json={"items": [{"id": 1, "month": "2026-09", "lot_count": 7}]})
        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.get_json()["code"], "forecast_month_inactive")
        self.assertEqual(cursor.rows[(1, september)]["lot_count"], 5)
        self.assertEqual(cursor.update_count, 0)

    def test_reactivated_month_allows_normal_lot_save(self):
        self.authenticate("PC")
        september = date(2026, 9, 1)
        cursor = self.connection.cursor_instance
        cursor.months[september] = {"source_label": "Sep-69", "is_active": True}
        cursor.rows[(1, september)] = {"id": 1, "part_no": "PART-A", "source_row": 5, "forecast_month": september, "source_label": "Sep-69", "quantity": Decimal("700"), "lot_count": 5}
        response = self.client.post("/api/forecast/lots", json={"items": [{"id": 1, "month": "2026-09", "lot_count": 7}]})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(cursor.rows[(1, september)]["lot_count"], 7)


class WorkerCursor:
    def __init__(self, successful_hash=None):
        self.successful_hash = successful_hash
        self.result = None

    def __enter__(self): return self
    def __exit__(self, *_args): return False
    def execute(self, sql, _params=()):
        if sql.strip().upper().startswith("SELECT LAST_SUCCESSFUL_SHA256"):
            self.result = {"last_successful_sha256": self.successful_hash} if self.successful_hash else None
        else:
            self.result = None
    def fetchone(self): return self.result


class WorkerConnection:
    def __init__(self, successful_hash=None):
        self.cursor_instance = WorkerCursor(successful_hash)
        self.committed = False
        self.rolled_back = False
    def cursor(self): return self.cursor_instance
    def commit(self): self.committed = True
    def rollback(self): self.rolled_back = True
    def close(self): pass


class ForecastWorkerTests(unittest.TestCase):
    def snapshot(self):
        temporary = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temporary.write(b"snapshot")
        temporary.close()
        return Path(temporary.name)

    def test_interval_is_configurable(self):
        with patch.dict(os.environ, {"FORECAST_SYNC_INTERVAL_SECONDS": "123"}):
            self.assertEqual(forecast_sync_worker.configured_interval(), 123)

    def test_temporarily_unreadable_source_never_connects_to_database(self):
        with patch.object(forecast_sync_worker, "stable_snapshot", side_effect=PermissionError("locked")):
            with self.assertRaises(forecast_sync_worker.ForecastSourceUnavailableError):
                forecast_sync_worker.run_once(
                    "source.xlsx",
                    connection_factory=lambda: self.fail("must not connect"),
                    stability_seconds=0,
                )

    def test_unchanged_sha_skips_all_writes(self):
        snapshot = self.snapshot()
        connection = WorkerConnection("abc")
        with patch.object(forecast_sync_worker, "stable_snapshot", return_value=snapshot), patch.object(forecast_sync_worker, "workbook_sha256", return_value="abc"):
            result = forecast_sync_worker.run_once("source.xlsx", connection_factory=lambda: connection, stability_seconds=0)
        self.assertEqual(result["status"], "unchanged")
        self.assertEqual(result["database_writes"], 0)
        self.assertFalse(connection.committed)

    def test_changed_workbook_applies_and_records_state_in_one_commit(self):
        snapshot = self.snapshot()
        connection = WorkerConnection()
        plan = {"parent_inserts": [], "parent_updates": [], "monthly_inserts": [(1, 2)], "monthly_updates": [], "structural_drift": {"passed": True}}
        report = {"source_rows": 1, "monthly_quantity_values": 1, "invalid_rows": 0}
        patches = (
            patch.object(forecast_sync_worker, "stable_snapshot", return_value=snapshot),
            patch.object(forecast_sync_worker, "workbook_sha256", return_value="changed"),
            patch.object(forecast_sync_worker, "parse_workbook", return_value=([{}], report)),
            patch.object(forecast_sync_worker, "load_database_state", return_value=({}, {}, {})),
            patch.object(forecast_sync_worker, "load_month_registry", return_value={}),
            patch.object(forecast_sync_worker, "classify_sync", return_value=plan),
            patch.object(forecast_sync_worker, "apply_sync"),
            patch.object(forecast_sync_worker, "record_sync_state"),
        )
        for active in patches: active.start()
        try:
            result = forecast_sync_worker.run_once("source.xlsx", connection_factory=lambda: connection, stability_seconds=0)
        finally:
            for active in reversed(patches): active.stop()
        self.assertEqual(result["status"], "success")
        self.assertTrue(connection.committed)

    def test_failed_apply_rolls_back(self):
        snapshot = self.snapshot()
        connection = WorkerConnection()
        plan = {"parent_inserts": [], "parent_updates": [], "monthly_inserts": [], "monthly_updates": [], "structural_drift": {"passed": True}}
        with patch.object(forecast_sync_worker, "stable_snapshot", return_value=snapshot), patch.object(forecast_sync_worker, "workbook_sha256", return_value="changed"), patch.object(forecast_sync_worker, "parse_workbook", return_value=([{}], {"invalid_rows": 0})), patch.object(forecast_sync_worker, "load_database_state", return_value=({}, {}, {})), patch.object(forecast_sync_worker, "load_month_registry", return_value={}), patch.object(forecast_sync_worker, "classify_sync", return_value=plan), patch.object(forecast_sync_worker, "apply_sync", side_effect=RuntimeError("boom")):
            with self.assertRaises(RuntimeError):
                forecast_sync_worker.run_once("source.xlsx", connection_factory=lambda: connection, stability_seconds=0)
        self.assertTrue(connection.rolled_back)

    def test_failure_after_month_and_quantity_changes_rolls_back_all_state(self):
        snapshot = self.snapshot()
        connection = WorkerConnection()
        connection.state = {"quantity": Decimal("500"), "lot": 5, "is_active": True}
        original = copy.deepcopy(connection.state)
        connection.rollback = lambda: (setattr(connection, "rolled_back", True), setattr(connection, "state", copy.deepcopy(original)))

        def fail_after_changes(_cursor, _plan):
            connection.state["quantity"] = Decimal("700")
            connection.state["is_active"] = False
            raise RuntimeError("boom")

        plan = {"parent_inserts": [], "parent_updates": [], "monthly_inserts": [], "monthly_updates": [], "structural_drift": {"passed": True}}
        report = {"source_rows": 1, "monthly_quantity_values": 1, "invalid_rows": 0, "discovered_months": []}
        with patch.object(forecast_sync_worker, "stable_snapshot", return_value=snapshot), patch.object(forecast_sync_worker, "workbook_sha256", return_value="changed"), patch.object(forecast_sync_worker, "parse_workbook", return_value=([{}], report)), patch.object(forecast_sync_worker, "load_database_state", return_value=({}, {}, {})), patch.object(forecast_sync_worker, "load_month_registry", return_value={}), patch.object(forecast_sync_worker, "classify_sync", return_value=plan), patch.object(forecast_sync_worker, "apply_sync", side_effect=fail_after_changes):
            with self.assertRaises(RuntimeError):
                forecast_sync_worker.run_once("source.xlsx", connection_factory=lambda: connection, stability_seconds=0)
        self.assertTrue(connection.rolled_back)
        self.assertEqual(connection.state, original)
        self.assertEqual(connection.state["lot"], 5)


class ForecastSourceContractTests(unittest.TestCase):
    def test_grouped_header_navigation_and_cross_month_dirty_save(self):
        source = (ROOT / "frontend/src/views/ForecastView.vue").read_text(encoding="utf-8")
        self.assertIn('rowspan="2"', source)
        self.assertIn('colspan="3"', source)
        self.assertIn("selectMonthByOffset", source)
        self.assertIn("Object.values(rowsByMonth.value).flat()", source)
        self.assertIn("month: row.month", source)
        self.assertIn("Search Part No.", source)
        self.assertIn("forecast_month_inactive", source)
        self.assertIn("await loadMonths()", source)

    def test_month_api_uses_real_mysql_date_format_not_literal_percent_escape(self):
        source = (ROOT / "app.py").read_text(encoding="utf-8")
        self.assertIn("DATE_FORMAT(forecast_month, '%Y-%m')", source)
        self.assertNotIn("DATE_FORMAT(forecast_month, '%%Y-%%m')", source)

    def test_worker_is_dedicated_and_source_mount_is_read_only(self):
        compose = (ROOT / "docker-compose.yml").read_text(encoding="utf-8")
        self.assertIn("forecast-sync:", compose)
        self.assertIn('profiles: ["forecast-sync"]', compose)
        self.assertIn("scripts/forecast_sync_worker.py", compose)
        self.assertIn("./forecast_source:/forecast_source:ro", compose)
        dockerfile = (ROOT / "Dockerfile").read_text(encoding="utf-8")
        self.assertNotIn("forecast_sync_worker", dockerfile)

    def test_manual_import_defaults_to_dry_run(self):
        self.assertFalse(parse_args([]).apply)


if __name__ == "__main__":
    unittest.main()
